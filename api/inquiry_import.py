"""Bulk-import an OP (outpatient) list into PRM inquiries.

A PRO uploads the day's OP list (CSV or .xlsx). Each row becomes an ``Inquiry``
with ``source=OP_IMPORT`` and ``status=NEW``. Rows are validated and imported
independently — a bad row is reported and skipped, the rest proceed.

Recognised columns (header row required; matched case-insensitively, with a few
aliases):

    name   (required)   — also: patient, patient name
    phone  (optional)   — also: mobile, contact, phone number
    notes  (optional)   — also: remarks, note

Dedup: re-uploading the same list is idempotent. A row's dedup key is its phone
if present, else its lowercased name. A row is skipped as a duplicate when an
existing OP-imported inquiry already has that key, or the same key appeared
earlier in the same file. Manually-entered inquiries (other sources) never
block an import.
"""
import csv
import io
from datetime import datetime

from .models import Inquiry, InquirySource, InquiryStatus

# Header aliases → canonical field. Keys are compared lowercased/stripped.
_ALIASES = {
    "name": "name",
    "patient": "name",
    "patient name": "name",
    "phone": "phone",
    "phone number": "phone",
    "mobile": "phone",
    "mobile number": "phone",
    "contact": "phone",
    "notes": "notes",
    "note": "notes",
    "remarks": "notes",
    "consult date": "consulted_on",
    "consulted on": "consulted_on",
    "consultation date": "consulted_on",
    "consult_date": "consulted_on",
    "op date": "consulted_on",
    "date": "consulted_on",
}

# Consult-date formats the OP list may use (Indian day-first, plus ISO and the
# "YYYY-MM-DD HH:MM:SS" that openpyxl stringifies date cells to).
_DATE_FORMATS = ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S")


def _parse_consult_date(value: str):
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None

_ALLOWED_EXTENSIONS = (".csv", ".xlsx")


class ImportFileError(Exception):
    """A whole-file problem (unreadable, wrong type, no header) — as opposed to
    a per-row error, which is collected and reported per row."""


def _canonical_headers(raw_headers):
    """Map a sheet's raw header cells to canonical field names, by position.
    Unknown headers map to None (their column is ignored)."""
    headers = []
    for h in raw_headers:
        key = str(h or "").strip().lower()
        headers.append(_ALIASES.get(key))
    return headers


def _rows_from_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ImportFileError("The file is empty.")
    headers = _canonical_headers(rows[0])
    return headers, rows[1:]


def _rows_from_xlsx(data: bytes):
    # Imported lazily so the dependency is only needed when an .xlsx is used.
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any parse failure as file error
        raise ImportFileError(f"Could not read the Excel file: {exc}")
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    # Drop trailing fully-empty rows (common in spreadsheets).
    while rows and all(c is None or str(c).strip() == "" for c in rows[-1]):
        rows.pop()
    if not rows:
        raise ImportFileError("The file is empty.")
    headers = _canonical_headers(rows[0])
    return headers, rows[1:]


def _parse(filename: str, data: bytes):
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _rows_from_csv(data)
    if name.endswith(".xlsx"):
        return _rows_from_xlsx(data)
    raise ImportFileError(
        "Unsupported file type. Upload a .csv or .xlsx OP list."
    )


def import_op_list(filename: str, data: bytes, user):
    """Parse and import an OP list. Returns a summary dict::

        {"total", "created", "duplicates", "errors": [{"row", "message"}]}

    ``total`` counts data rows considered (excludes the header and blank rows).
    Raises ``ImportFileError`` for whole-file problems.
    """
    headers, data_rows = _parse(filename, data)
    if "name" not in headers:
        raise ImportFileError(
            "Missing a 'name' column. Recognised columns: name, phone, notes."
        )

    # Existing OP-import dedup keys (phone preferred, else lowercased name).
    seen = set()
    for inq in Inquiry.objects.filter(source=InquirySource.OP_IMPORT).values(
        "name", "phone"
    ):
        key = (inq["phone"] or "").strip() or (inq["name"] or "").strip().lower()
        if key:
            seen.add(key)

    created = 0
    duplicates = 0
    errors = []
    to_create = []

    for offset, raw in enumerate(data_rows, start=2):  # header is row 1
        # Map this row's cells to fields by header position.
        record = {"name": "", "phone": "", "notes": "", "consulted_on": ""}
        for col, field in enumerate(headers):
            if field is None or col >= len(raw):
                continue
            record[field] = str(raw[col] if raw[col] is not None else "").strip()

        # Skip a fully-blank row silently (not an error, not counted).
        if not any(record.values()):
            continue

        if not record["name"]:
            errors.append({"row": offset, "message": "name is required"})
            continue

        consulted_on = None
        if record["consulted_on"]:
            consulted_on = _parse_consult_date(record["consulted_on"])
            if consulted_on is None:
                errors.append({"row": offset, "message": "consult date must be DD-MM-YYYY"})
                continue

        key = record["phone"] or record["name"].lower()
        if key in seen:
            duplicates += 1
            continue
        seen.add(key)

        to_create.append(
            Inquiry(
                name=record["name"],
                phone=record["phone"][:20],
                notes=record["notes"],
                consulted_on=consulted_on,
                source=InquirySource.OP_IMPORT,
                status=InquiryStatus.NEW,
                created_by=user,
            )
        )
        created += 1

    if to_create:
        Inquiry.objects.bulk_create(to_create)

    total = created + duplicates + len(errors)
    return {
        "total": total,
        "created": created,
        "duplicates": duplicates,
        "errors": errors,
    }
